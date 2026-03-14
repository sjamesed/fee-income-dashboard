"""Excel file parser for fee income data."""

import re
import logging
from pathlib import Path
from collections import defaultdict

import openpyxl

logger = logging.getLogger(__name__)

MONTH_MAP = {
    "jan": "01", "feb": "02", "mar": "03", "apr": "04",
    "may": "05", "jun": "06", "jul": "07", "aug": "08",
    "sep": "09", "oct": "10", "nov": "11", "dec": "12",
}

KNOWN_FEE_TYPES = {
    "Asset Mgmt Fee", "Acq / Div Fee", "Development Mgmt Fee",
    "Leasing Fee", "Other Fee", "Promote Fee",
}

SHEET_NAME = "Summary_new template (작성탭)"
HEADER_ROW = 20
DATA_START_ROW = 21
EMPTY_ROW_THRESHOLD = 10


def extract_snapshot_from_filename(filename: str) -> str:
    match = re.search(r"\((\d+\+\d+)\)", filename)
    if not match:
        raise ValueError(f"Cannot extract snapshot from filename: {filename}")
    return match.group(1)


def parse_annual_header(header: str) -> tuple[str, str]:
    header = header.strip()
    period_match = re.match(r"((?:FY|[12]H)\d{2})", header)
    if not period_match:
        raise ValueError(f"Cannot parse annual header: {header}")
    period = period_match.group(1)
    remainder = header[period_match.end():].strip().lower()
    remainder = remainder.strip("()")
    if "refcst" in remainder or "reforecast" in remainder or "5yr" in remainder:
        return period, "reforecast"
    elif "bud" in remainder:
        return period, "budget"
    elif "fcst" in remainder or "forecast" in remainder:
        return period, "forecast"
    elif "act" in remainder or remainder == "":
        return period, "actual"
    else:
        logger.warning(f"Unknown annual header qualifier: '{header}', defaulting to actual")
        return period, "actual"


def parse_monthly_header(header: str) -> tuple[str, str]:
    header = header.strip()
    match = re.match(r"([A-Za-z]+)[\s\-]+(\d{2})\s*(?:\((.+?)\))?$", header)
    if not match:
        raise ValueError(f"Cannot parse monthly header: {header}")
    month_str, year_str, qualifier = match.group(1), match.group(2), match.group(3)
    month_num = MONTH_MAP.get(month_str.lower())
    if not month_num:
        raise ValueError(f"Unknown month: {month_str} in header: {header}")
    year = f"20{year_str}"
    period = f"{year}-{month_num}"
    if qualifier is None:
        return period, "actual"
    qualifier = qualifier.strip().lower()
    if qualifier == "act":
        return period, "actual"
    elif qualifier == "bud":
        return period, "budget"
    elif re.match(r"\d+\+\d+", qualifier):
        return period, "forecast"
    else:
        logger.warning(f"Unknown monthly qualifier: '{qualifier}' in header: {header}")
        return period, "actual"


def _validate_fy_cross_check(rows: list[dict]):
    monthly_sums = defaultdict(float)
    fy_totals = defaultdict(float)
    for r in rows:
        key = (r["platform"], r["project_name"], r["fee_type"])
        period = r["period"]
        if r["period_type"] == "actual" and re.match(r"\d{4}-\d{2}", period):
            year = period[:4]
            fy_key = (*key, f"FY{year[2:]}")
            monthly_sums[fy_key] += r["amount_usd"]
        elif r["period_type"] == "actual" and period.startswith("FY"):
            fy_key = (*key, period)
            fy_totals[fy_key] = r["amount_usd"]
    for fy_key, fy_total in fy_totals.items():
        monthly_sum = monthly_sums.get(fy_key, 0)
        if fy_total != 0:
            diff_pct = abs(fy_total - monthly_sum) / abs(fy_total) * 100
            if diff_pct > 1:
                platform, project, fee_type, period = fy_key
                logger.warning(
                    f"FY cross-check: {project}/{fee_type}/{period} — "
                    f"FY total={fy_total:.0f}, monthly sum={monthly_sum:.0f}, diff={diff_pct:.1f}%"
                )


def parse_excel_file(filepath: str) -> tuple[list[dict], str]:
    filename = Path(filepath).name
    snapshot = extract_snapshot_from_filename(filename)
    wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
    if SHEET_NAME not in wb.sheetnames:
        available = ", ".join(wb.sheetnames[:10])
        wb.close()
        raise ValueError(f"Sheet '{SHEET_NAME}' not found. Available sheets: {available}...")
    ws = wb[SHEET_NAME]

    # Build column mapping from headers
    header_cells = {}
    for row in ws.iter_rows(min_row=HEADER_ROW, max_row=HEADER_ROW, max_col=200):
        for cell in row:
            if cell.value is not None:
                header_cells[cell.column] = cell.value

    col_map = {}
    for col_idx, header_val in header_cells.items():
        if col_idx <= 6:
            continue
        header = str(header_val).strip()
        if not header:
            continue
        try:
            if re.match(r"(?:FY|[12]H)\d{2}", header):
                period, period_type = parse_annual_header(header)
            else:
                period, period_type = parse_monthly_header(header)
            col_map[col_idx] = (period, period_type)
        except ValueError as e:
            logger.warning(f"Skipping column {col_idx} with header '{header}': {e}")

    max_col = max(col_map.keys()) if col_map else 200
    rows = []
    empty_streak = 0

    for row in ws.iter_rows(min_row=DATA_START_ROW, max_col=max_col):
        row_dict = {cell.column: cell.value for cell in row}
        platform = row_dict.get(2)
        project_name = row_dict.get(3)
        project_status = row_dict.get(4)
        risk_category = row_dict.get(5)
        fee_type = row_dict.get(6)

        dims_empty = all(row_dict.get(c) is None for c in range(2, 7))
        if dims_empty:
            empty_streak += 1
            if empty_streak >= EMPTY_ROW_THRESHOLD:
                break
            continue
        empty_streak = 0

        if not platform:
            continue
        if fee_type and fee_type not in KNOWN_FEE_TYPES:
            if fee_type in ("Committed", "Committed w/ risk", "Uncommitted"):
                continue
            logger.warning(f"Unknown fee type: '{fee_type}' for {project_name}")
        if not fee_type:
            continue

        for col_idx, (period, period_type) in col_map.items():
            amount = row_dict.get(col_idx)
            if amount is None:
                amount = 0.0
            try:
                amount = float(amount)
            except (ValueError, TypeError):
                amount = 0.0
            rows.append({
                "snapshot": snapshot,
                "platform": str(platform),
                "project_name": str(project_name),
                "project_status": str(project_status) if project_status else None,
                "risk_category": str(risk_category) if risk_category else None,
                "fee_type": str(fee_type),
                "period_type": period_type,
                "period": period,
                "amount_usd": amount,
            })

    wb.close()
    _validate_fy_cross_check(rows)
    logger.info(f"Parsed {len(rows)} data points from {filename} (snapshot: {snapshot})")
    return rows, snapshot
