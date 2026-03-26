"""
Parser for capital_call.xlsx -> capital_call.json

Extracts per-fund, per-project capital call data from each sheet.
For sheets with ESR/CPPIB split (Terra), focuses on ESR's portion.
For other sheets, takes total amounts.

Produces both KRW and USD sections where available.
"""

import json
import os
from datetime import datetime
from pathlib import Path

import openpyxl

BASE_DIR = Path(__file__).resolve().parent.parent
EXCEL_PATH = BASE_DIR / "data" / "capital_call.xlsx"
JSON_PATH = BASE_DIR / "data" / "capital_call.json"


def safe_float(val):
    """Convert a value to float, returning None if not numeric."""
    if val is None:
        return None
    try:
        return float(val)
    except (ValueError, TypeError):
        return None


def safe_date(val):
    """Convert a value to ISO date string."""
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d")
    return str(val)


def parse_terra(wb):
    """
    Capital call-Terra sheet.
    Row 2: headers - project names at cols 6,8,10,12,14,16,18,20,23,25
    Row 3: ESR/CPPIB split - ESR cols: 6,8,10,12,14,16,18,20,25; CPPIB: 7,9,11,13,15,17,19,21,23
    Data rows: 4 onwards (KRW section), 27 onwards (USD section)

    Focus on ESR portion (col 6=Total ESR, project ESR cols: 8,10,12,14,16,18,20,25)
    For Juangan (col 23) - only CPPIB column exists, skip ESR.
    For Maplewood (col 25) - ESR column.
    """
    ws = wb["Capital call-Terra"]

    # Project mapping: project_name -> (esr_krw_col, esr_usd_col)
    # KRW section rows 2-3, USD section starts at row 25
    # ESR columns in KRW: 8(NHN), 10(Mokcheon), 12(DSL), 14(Fila), 16(Assist Korea), 18(AMB Icheon), 20(Sangwoo Logis), 25(Maplewood)
    # Juangan only has CPPIB (col 23), no ESR column
    projects_krw = {
        "NHN": 8,
        "Mokcheon": 10,
        "DSL": 12,
        "Fila": 14,
        "Assist Korea": 16,
        "AMB Icheon": 18,
        "Sangwoo Logis": 20,
        "Maplewood": 25,
    }

    projects_usd = {
        "NHN": 8,
        "Mokcheon": 10,
        "DSL": 12,
        "Fila": 14,
        "Assist Korea": 16,
        "AMB Icheon": 18,
        "Sangwoo Logis": 20,
        "Maplewood": 25,
    }

    # KRW data rows: 4 to 13 (row 14 is Sum)
    krw_data_end = 13
    # USD data rows: 27 to 36 (row 37 is Sum)
    usd_data_start = 27
    usd_data_end = 36

    # Build per-project call lists
    project_calls = {name: [] for name in projects_krw}

    # Parse KRW rows
    for row in range(4, krw_data_end + 1):
        date_val = safe_date(ws.cell(row, 3).value)
        fx_rate = safe_float(ws.cell(row, 4).value)
        details = ws.cell(row, 5).value
        if details is None:
            details = ""
        details = str(details).strip()

        # If no date on this row, inherit from the previous call row
        # (continuation rows like Leasing Fee, Acquisition Fee)
        if date_val is None:
            # Look backwards for the nearest date
            for prev_row in range(row - 1, 3, -1):
                d = ws.cell(prev_row, 3).value
                if d is not None:
                    date_val = safe_date(d)
                    if fx_rate is None:
                        fx_rate = safe_float(ws.cell(prev_row, 4).value)
                    break

        for proj_name, col in projects_krw.items():
            krw_val = safe_float(ws.cell(row, col).value)
            if krw_val is not None and krw_val != 0:
                project_calls[proj_name].append({
                    "date": date_val,
                    "krw": krw_val,
                    "usd": None,
                    "fx_rate": fx_rate,
                    "details": details,
                })

    # Parse USD rows and match to existing calls or add USD info
    usd_by_project = {name: [] for name in projects_usd}
    for row in range(usd_data_start, usd_data_end + 1):
        date_val = safe_date(ws.cell(row, 3).value)
        fx_rate = safe_float(ws.cell(row, 4).value)
        details = ws.cell(row, 5).value
        if details is None:
            details = ""
        details = str(details).strip()

        if date_val is None:
            for prev_row in range(row - 1, usd_data_start - 1, -1):
                d = ws.cell(prev_row, 3).value
                if d is not None:
                    date_val = safe_date(d)
                    if fx_rate is None:
                        fx_rate = safe_float(ws.cell(prev_row, 4).value)
                    break

        for proj_name, col in projects_usd.items():
            usd_val = safe_float(ws.cell(row, col).value)
            if usd_val is not None and usd_val != 0:
                usd_by_project[proj_name].append({
                    "date": date_val,
                    "usd": usd_val,
                    "details": details,
                })

    # Merge USD into KRW calls where dates+details match, or add as separate
    for proj_name in projects_krw:
        for usd_entry in usd_by_project.get(proj_name, []):
            matched = False
            for call in project_calls[proj_name]:
                if call["date"] == usd_entry["date"] and call["usd"] is None:
                    call["usd"] = usd_entry["usd"]
                    matched = True
                    break
            if not matched:
                # USD-only call (like Maplewood entries)
                project_calls[proj_name].append({
                    "date": usd_entry["date"],
                    "krw": None,
                    "usd": usd_entry["usd"],
                    "fx_rate": None,
                    "details": usd_entry["details"],
                })

    # Build result
    projects = []
    for proj_name in projects_krw:
        calls = project_calls[proj_name]
        if not calls:
            continue
        total_krw = sum(c["krw"] for c in calls if c["krw"] is not None)
        total_usd = sum(c["usd"] for c in calls if c["usd"] is not None)
        # Sort by date
        calls.sort(key=lambda c: c["date"] or "")
        projects.append({
            "project": proj_name,
            "calls": calls,
            "total_krw": total_krw if total_krw else None,
            "total_usd": total_usd if total_usd else None,
        })

    return {
        "fund": "Terra",
        "sheet": "Capital call-Terra",
        "projects": projects,
    }


def parse_acqui_fund(wb):
    """
    Acqui fund sheet. Simple table.
    Row 2: headers. Data rows: 3 onwards.
    Col A=fund#, B=project, C=request date, D=wired date, E=KRW, F=FX rate, G=USD.
    """
    ws = wb["Acqui fund"]
    projects = {}

    for row in range(3, ws.max_row + 1):
        fund_num = ws.cell(row, 1).value
        project = ws.cell(row, 2).value
        if project is None:
            continue
        project = str(project).strip()

        req_date = safe_date(ws.cell(row, 3).value)
        wired_date = safe_date(ws.cell(row, 4).value)
        krw = safe_float(ws.cell(row, 5).value)
        fx_rate = safe_float(ws.cell(row, 6).value)
        usd = safe_float(ws.cell(row, 7).value)

        if project not in projects:
            projects[project] = []

        details = f"Fund {fund_num}" if fund_num else ""
        projects[project].append({
            "date": wired_date or req_date,
            "krw": krw,
            "usd": usd,
            "fx_rate": fx_rate,
            "details": details,
        })

    result_projects = []
    for proj_name, calls in projects.items():
        total_krw = sum(c["krw"] for c in calls if c["krw"] is not None)
        total_usd = sum(c["usd"] for c in calls if c["usd"] is not None)
        calls.sort(key=lambda c: c["date"] or "")
        result_projects.append({
            "project": proj_name,
            "calls": calls,
            "total_krw": total_krw if total_krw else None,
            "total_usd": total_usd if total_usd else None,
        })

    return {
        "fund": "Acqui Fund",
        "sheet": "Acqui fund",
        "projects": result_projects,
    }


def parse_dev_jv_sheet(wb, sheet_name, fund_name, krw_header_row, krw_data_start,
                       krw_data_end, usd_header_row, usd_data_start, usd_data_end,
                       project_col_start, project_col_end,
                       date_col=3, fx_col=4, details_col=5):
    """
    Generic parser for Dev JV1 (Star), Dev JV2 (Nova), Sunwood Byul sheets.
    These have similar structure: project names in header row, amounts in columns below.
    One column per project (no ESR/CPPIB split - these are 100% ESR).
    """
    ws = wb[sheet_name]

    # Read project headers
    project_cols = {}  # project_name -> col
    for col in range(project_col_start, project_col_end + 1):
        name = ws.cell(krw_header_row, col).value
        if name is not None:
            name = str(name).strip()
            if name and name not in ("in KRW", "Value date", "Ex rate", "Details",
                                     "Capital Drawdown", "in USD"):
                project_cols[name] = col

    # Parse KRW data
    project_calls = {name: [] for name in project_cols}

    current_call_label = ""
    for row in range(krw_data_start, krw_data_end + 1):
        # Col B (col 2) often has "Capital Call 1", "Capital Call #2", etc.
        label_val = ws.cell(row, 2).value
        if label_val is not None:
            label_str = str(label_val).strip()
            if label_str.lower().startswith("capital call") or label_str.startswith("#"):
                current_call_label = label_str

        date_val = safe_date(ws.cell(row, date_col).value)
        fx_rate = safe_float(ws.cell(row, fx_col).value)
        details = ws.cell(row, details_col).value
        if details is None:
            details = ""
        details = str(details).strip()

        if date_val is None:
            for prev_row in range(row - 1, krw_data_start - 1, -1):
                d = ws.cell(prev_row, date_col).value
                if d is not None:
                    date_val = safe_date(d)
                    if fx_rate is None:
                        fx_rate = safe_float(ws.cell(prev_row, fx_col).value)
                    break

        for proj_name, col in project_cols.items():
            krw_val = safe_float(ws.cell(row, col).value)
            if krw_val is not None and krw_val != 0:
                project_calls[proj_name].append({
                    "date": date_val,
                    "krw": krw_val,
                    "usd": None,
                    "fx_rate": fx_rate,
                    "details": details,
                    "call_label": current_call_label,
                })

    # Parse USD data if available
    if usd_header_row and usd_data_start and usd_data_end:
        usd_project_cols = {}
        for col in range(project_col_start, project_col_end + 1):
            name = ws.cell(usd_header_row, col).value
            if name is not None:
                name = str(name).strip()
                if name and name not in ("in USD", "Value date", "Ex rate", "Details",
                                         "Capital Drawdown"):
                    usd_project_cols[name] = col

        # Match USD project names to KRW project names
        # They should be identical or very similar
        usd_to_krw_name = {}
        for usd_name in usd_project_cols:
            if usd_name in project_cols:
                usd_to_krw_name[usd_name] = usd_name
            else:
                # Try fuzzy match
                for krw_name in project_cols:
                    if usd_name.lower().replace(" ", "") == krw_name.lower().replace(" ", ""):
                        usd_to_krw_name[usd_name] = krw_name
                        break

        for row in range(usd_data_start, usd_data_end + 1):
            date_val = safe_date(ws.cell(row, date_col).value)
            fx_rate = safe_float(ws.cell(row, fx_col).value)
            details = ws.cell(row, details_col).value
            if details is None:
                details = ""
            details = str(details).strip()

            if date_val is None:
                for prev_row in range(row - 1, usd_data_start - 1, -1):
                    d = ws.cell(prev_row, date_col).value
                    if d is not None:
                        date_val = safe_date(d)
                        if fx_rate is None:
                            fx_rate = safe_float(ws.cell(prev_row, fx_col).value)
                        break

            for usd_proj, col in usd_project_cols.items():
                krw_proj = usd_to_krw_name.get(usd_proj)
                if krw_proj is None:
                    continue
                usd_val = safe_float(ws.cell(row, col).value)
                if usd_val is not None and usd_val != 0:
                    # Try to match to a KRW call with same date
                    matched = False
                    for call in project_calls.get(krw_proj, []):
                        if call["date"] == date_val and call["usd"] is None:
                            call["usd"] = usd_val
                            matched = True
                            break
                    if not matched:
                        if krw_proj not in project_calls:
                            project_calls[krw_proj] = []
                        project_calls[krw_proj].append({
                            "date": date_val,
                            "krw": None,
                            "usd": usd_val,
                            "fx_rate": fx_rate,
                            "details": details,
                        })

    # Build result
    projects = []
    for proj_name in project_cols:
        calls = project_calls.get(proj_name, [])
        if not calls:
            continue
        total_krw = sum(c["krw"] for c in calls if c["krw"] is not None)
        total_usd = sum(c["usd"] for c in calls if c["usd"] is not None)
        calls.sort(key=lambda c: c["date"] or "")
        projects.append({
            "project": proj_name,
            "calls": calls,
            "total_krw": total_krw if total_krw else None,
            "total_usd": total_usd if total_usd else None,
        })

    return {
        "fund": fund_name,
        "sheet": sheet_name,
        "projects": projects,
    }


def find_krw_data_end(ws, start_row, label_col=2):
    """Find the last data row before a 'Sum' row or empty section."""
    for row in range(start_row, ws.max_row + 1):
        val = ws.cell(row, label_col).value
        if val is not None and "sum" in str(val).lower().strip():
            return row - 1

    # If no Sum found, scan for gaps.
    # A data row should have at least one of: label (col 2), date (col 3), or details (col 5).
    # A row with only numeric values in project columns (col 6+) but nothing in cols 2-5
    # followed by empty rows is a sum/total row.
    last_data_row = start_row - 1
    for row in range(start_row, ws.max_row + 1):
        has_key_info = (
            ws.cell(row, 2).value is not None
            or ws.cell(row, 3).value is not None
            or ws.cell(row, 4).value is not None
            or ws.cell(row, 5).value is not None
        )
        has_any_data = False
        for col in range(2, 30):
            if ws.cell(row, col).value is not None:
                has_any_data = True
                break

        if has_key_info:
            last_data_row = row
        elif has_any_data:
            # Row with only project values, no key info - could be sum or continuation.
            # Check if next 2 rows are completely empty
            next_empty = 0
            for nr in range(row + 1, min(row + 4, ws.max_row + 1)):
                row_empty = True
                for col in range(2, 30):
                    if ws.cell(nr, col).value is not None:
                        row_empty = False
                        break
                if row_empty:
                    next_empty += 1
                else:
                    break
            if next_empty >= 2:
                # This is likely a sum row followed by empty space
                return last_data_row
            else:
                # Continuation data row (like row 6 in JV1 with amounts but no label)
                last_data_row = row
        else:
            # Completely empty row
            empty_count = 1
            for nr in range(row + 1, min(row + 4, ws.max_row + 1)):
                row_empty = True
                for col in range(2, 30):
                    if ws.cell(nr, col).value is not None:
                        row_empty = False
                        break
                if row_empty:
                    empty_count += 1
                else:
                    break
            if empty_count >= 3:
                return last_data_row

    return last_data_row if last_data_row >= start_row else ws.max_row


def parse_dev_jv1(wb):
    """Dev JV1_Capital call-Star"""
    ws = wb["Dev JV1_Capital call-Star"]
    # KRW: header row 2, data 3-169 (row 173 is Sum, but rows 170-172 may be empty)
    # Find actual end
    krw_end = find_krw_data_end(ws, 3, label_col=2)
    # USD: header row 183, data 184 to ~353
    usd_end = 353  # Last data row before summary

    return parse_dev_jv_sheet(
        wb,
        sheet_name="Dev JV1_Capital call-Star",
        fund_name="Dev JV1 (Star)",
        krw_header_row=2,
        krw_data_start=3,
        krw_data_end=krw_end,
        usd_header_row=183,
        usd_data_start=184,
        usd_data_end=usd_end,
        project_col_start=6,
        project_col_end=25,
    )


def parse_dev_jv2(wb):
    """Dev JV2_Capital call-Nova"""
    ws = wb["Dev JV2_Capital call-Nova"]
    krw_end = find_krw_data_end(ws, 3, label_col=2)

    # Find USD section
    usd_header = None
    for row in range(krw_end + 1, ws.max_row + 1):
        val = ws.cell(row, 2).value
        if val and "in USD" in str(val):
            usd_header = row
            break

    usd_start = usd_header + 1 if usd_header else None
    usd_end = None
    if usd_start:
        usd_end = find_krw_data_end(ws, usd_start, label_col=2)

    return parse_dev_jv_sheet(
        wb,
        sheet_name="Dev JV2_Capital call-Nova",
        fund_name="Dev JV2 (Nova)",
        krw_header_row=2,
        krw_data_start=3,
        krw_data_end=krw_end,
        usd_header_row=usd_header,
        usd_data_start=usd_start,
        usd_data_end=usd_end,
        project_col_start=6,
        project_col_end=14,
    )


def parse_sunwood_byul(wb):
    """Sunwood Byul sheet"""
    ws = wb["Sunwood Byul"]
    krw_end = find_krw_data_end(ws, 3, label_col=2)

    # Find USD section
    usd_header = None
    for row in range(krw_end + 1, ws.max_row + 1):
        val = ws.cell(row, 2).value
        if val and "in USD" in str(val):
            usd_header = row
            break

    usd_start = usd_header + 1 if usd_header else None
    usd_end = None
    if usd_start:
        usd_end = find_krw_data_end(ws, usd_start, label_col=2)

    return parse_dev_jv_sheet(
        wb,
        sheet_name="Sunwood Byul",
        fund_name="Sunwood Byul",
        krw_header_row=2,
        krw_data_start=3,
        krw_data_end=krw_end,
        usd_header_row=usd_header,
        usd_data_start=usd_start,
        usd_data_end=usd_end,
        project_col_start=6,
        project_col_end=9,
    )


def parse_dangmok(wb):
    """Dangmok sheet - single project"""
    ws = wb["Dangmok"]
    krw_end = find_krw_data_end(ws, 3, label_col=2)

    # Find USD section
    usd_header = None
    for row in range(krw_end + 1, ws.max_row + 1):
        val = ws.cell(row, 2).value
        if val and "in USD" in str(val):
            usd_header = row
            break

    usd_start = usd_header + 1 if usd_header else None
    usd_end = None
    if usd_start:
        usd_end = find_krw_data_end(ws, usd_start, label_col=2)

    return parse_dev_jv_sheet(
        wb,
        sheet_name="Dangmok",
        fund_name="Dangmok",
        krw_header_row=2,
        krw_data_start=3,
        krw_data_end=krw_end,
        usd_header_row=usd_header,
        usd_data_start=usd_start,
        usd_data_end=usd_end,
        project_col_start=6,
        project_col_end=6,
    )


def parse_income_jv(wb):
    """
    Income JV call sheet.
    Row 6: headers - Call#, Date, Detail, CPPIB, ESR
    Row 7: sub-headers - CPPIB REF, Sunwood Terra, KSAMC, Maplewood
    Data rows: 8 onwards.

    ESR portion is col 5 (Sunwood Terra).
    Also has KSAMC (col 6) and Maplewood (col 7).
    """
    ws = wb["Income JV call"]

    calls = []
    for row in range(8, ws.max_row + 1):
        call_num = ws.cell(row, 1).value
        date_val = safe_date(ws.cell(row, 2).value)
        details = ws.cell(row, 3).value
        if details is None:
            details = ""
        details = str(details).strip()

        if details.lower().startswith("total"):
            break

        # Skip non-data rows
        if call_num is None and date_val is None:
            continue

        # ESR (Sunwood Terra) = col 5
        esr_krw = safe_float(ws.cell(row, 5).value)
        # CPPIB = col 4
        cppib_krw = safe_float(ws.cell(row, 4).value)
        # KSAMC = col 6
        ksamc_krw = safe_float(ws.cell(row, 6).value)
        # Maplewood = col 7
        maple_krw = safe_float(ws.cell(row, 7).value)

        call_label = str(call_num) if call_num else ""

        if date_val is None:
            for prev_row in range(row - 1, 7, -1):
                d = ws.cell(prev_row, 2).value
                if d is not None:
                    date_val = safe_date(d)
                    break

        entry = {
            "date": date_val,
            "call_num": call_label,
            "details": details,
            "esr_krw": esr_krw,
            "cppib_krw": cppib_krw,
            "ksamc_krw": ksamc_krw,
            "maplewood_krw": maple_krw,
        }
        calls.append(entry)

    # Build per-entity project entries
    projects = []

    # ESR (Sunwood Terra)
    esr_calls = []
    for c in calls:
        if c["esr_krw"] is not None and c["esr_krw"] != 0:
            esr_calls.append({
                "date": c["date"],
                "krw": c["esr_krw"],
                "usd": None,
                "fx_rate": None,
                "details": c["details"],
            })
    if esr_calls:
        total = sum(x["krw"] for x in esr_calls if x["krw"])
        projects.append({
            "project": "ESR (Sunwood Terra)",
            "calls": esr_calls,
            "total_krw": total,
            "total_usd": None,
        })

    # CPPIB
    cppib_calls = []
    for c in calls:
        if c["cppib_krw"] is not None and c["cppib_krw"] != 0:
            cppib_calls.append({
                "date": c["date"],
                "krw": c["cppib_krw"],
                "usd": None,
                "fx_rate": None,
                "details": c["details"],
            })
    if cppib_calls:
        total = sum(x["krw"] for x in cppib_calls if x["krw"])
        projects.append({
            "project": "CPPIB",
            "calls": cppib_calls,
            "total_krw": total,
            "total_usd": None,
        })

    # KSAMC
    ksamc_calls = []
    for c in calls:
        if c["ksamc_krw"] is not None and c["ksamc_krw"] != 0:
            ksamc_calls.append({
                "date": c["date"],
                "krw": c["ksamc_krw"],
                "usd": None,
                "fx_rate": None,
                "details": c["details"],
            })
    if ksamc_calls:
        total = sum(x["krw"] for x in ksamc_calls if x["krw"])
        projects.append({
            "project": "KSAMC",
            "calls": ksamc_calls,
            "total_krw": total,
            "total_usd": None,
        })

    # Maplewood
    maple_calls = []
    for c in calls:
        if c["maplewood_krw"] is not None and c["maplewood_krw"] != 0:
            maple_calls.append({
                "date": c["date"],
                "krw": c["maplewood_krw"],
                "usd": None,
                "fx_rate": None,
                "details": c["details"],
            })
    if maple_calls:
        total = sum(x["krw"] for x in maple_calls if x["krw"])
        projects.append({
            "project": "Maplewood",
            "calls": maple_calls,
            "total_krw": total,
            "total_usd": None,
        })

    return {
        "fund": "Income JV",
        "sheet": "Income JV call",
        "projects": projects,
    }


def main():
    print(f"Loading workbook: {EXCEL_PATH}")
    wb = openpyxl.load_workbook(str(EXCEL_PATH), data_only=True)
    print(f"Sheets: {wb.sheetnames}")

    funds = []

    # 1. Terra
    print("Parsing: Capital call-Terra...")
    terra = parse_terra(wb)
    funds.append(terra)
    print(f"  -> {len(terra['projects'])} projects, "
          f"{sum(len(p['calls']) for p in terra['projects'])} calls")

    # 2. Acqui Fund
    print("Parsing: Acqui fund...")
    acqui = parse_acqui_fund(wb)
    funds.append(acqui)
    print(f"  -> {len(acqui['projects'])} projects, "
          f"{sum(len(p['calls']) for p in acqui['projects'])} calls")

    # 3. Dev JV1 (Star)
    print("Parsing: Dev JV1_Capital call-Star...")
    jv1 = parse_dev_jv1(wb)
    funds.append(jv1)
    print(f"  -> {len(jv1['projects'])} projects, "
          f"{sum(len(p['calls']) for p in jv1['projects'])} calls")

    # 4. Dev JV2 (Nova)
    print("Parsing: Dev JV2_Capital call-Nova...")
    jv2 = parse_dev_jv2(wb)
    funds.append(jv2)
    print(f"  -> {len(jv2['projects'])} projects, "
          f"{sum(len(p['calls']) for p in jv2['projects'])} calls")

    # 5. Sunwood Byul
    print("Parsing: Sunwood Byul...")
    byul = parse_sunwood_byul(wb)
    funds.append(byul)
    print(f"  -> {len(byul['projects'])} projects, "
          f"{sum(len(p['calls']) for p in byul['projects'])} calls")

    # 6. Dangmok
    print("Parsing: Dangmok...")
    dangmok = parse_dangmok(wb)
    funds.append(dangmok)
    print(f"  -> {len(dangmok['projects'])} projects, "
          f"{sum(len(p['calls']) for p in dangmok['projects'])} calls")

    # 7. Income JV
    print("Parsing: Income JV call...")
    income = parse_income_jv(wb)
    funds.append(income)
    print(f"  -> {len(income['projects'])} projects, "
          f"{sum(len(p['calls']) for p in income['projects'])} calls")

    # Summary
    result = {"funds": funds}

    print(f"\nWriting JSON to: {JSON_PATH}")
    with open(JSON_PATH, "w", encoding="utf-8") as f:
        json.dump(result, f, indent=2, ensure_ascii=False)

    # Print summary
    total_projects = sum(len(fund["projects"]) for fund in funds)
    total_calls = sum(
        sum(len(p["calls"]) for p in fund["projects"])
        for fund in funds
    )
    print(f"\nDone! {len(funds)} funds, {total_projects} projects, {total_calls} total calls.")

    # Print per-fund totals
    for fund in funds:
        print(f"\n  {fund['fund']}:")
        for proj in fund["projects"]:
            krw_str = f"{proj['total_krw']:,.0f}" if proj["total_krw"] else "N/A"
            usd_str = f"{proj['total_usd']:,.0f}" if proj["total_usd"] else "N/A"
            print(f"    {proj['project']}: KRW {krw_str} / USD {usd_str} ({len(proj['calls'])} calls)")


if __name__ == "__main__":
    main()
