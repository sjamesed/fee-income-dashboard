"""Page 5: Data Management — upload, view, and manage snapshot data."""
import streamlit as st
import pandas as pd
import tempfile
import os
import json
from pathlib import Path
from src.db import FeeIncomeDB
from src.parser import parse_excel_file

DISPOSAL_DIR = Path(__file__).parent.parent / "data"


def extract_mm_report_json(xlsx_bytes, snapshot_name):
    """Extract Output PL + Output SG&A + 1b. PL-breakdown data from MM Report and save as JSON."""
    import openpyxl
    tmp_path = DISPOSAL_DIR / f"_tmp_mm_{snapshot_name}.xlsx"
    tmp_path.write_bytes(xlsx_bytes)
    try:
        wb = openpyxl.load_workbook(str(tmp_path), data_only=True, read_only=True)

        result = {}

        # --- Output PL ---
        ws = wb["Output PL"]
        PL_ROWS = [5,7,8,9,10,11,12,13,14,15,17,18,19,20,21,22,23,24,25,
                   27,28,29,30,31,33,34,36,37,39,40,42,43,44,46,47,49,50,51,52,53,54,56,58,60]
        PL_COLS = [2, 4,5,6,7, 8,9,10,11, 26,27,28,29, 30,31,32,33]  # 1-based
        pl_data = {}
        for r in PL_ROWS:
            rd = {}
            for c in PL_COLS:
                v = ws.cell(row=r, column=c).value
                if v is not None:
                    rd[str(c)] = v if isinstance(v, (int, float)) else str(v)
            pl_data[str(r)] = rd
        result["output_pl"] = pl_data

        # --- Output SG&A ---
        ws2 = wb["Output SG&A"]
        SGA_ROWS = list(range(3, 21))
        SGA_COLS = list(range(1, 28))
        sga_data = {}
        for r in SGA_ROWS:
            rd = {}
            for c in SGA_COLS:
                v = ws2.cell(row=r, column=c).value
                if v is not None:
                    rd[str(c)] = v if isinstance(v, (int, float)) else str(v)
            if rd:
                sga_data[str(r)] = rd
        result["output_sga"] = sga_data

        # --- 1b. PL-breakdown ---
        ws3 = wb["1b. PL-breakdown"]
        bd_data = []
        for r in range(4, ws3.max_row + 1):
            g = ws3.cell(row=r, column=7).value
            h = ws3.cell(row=r, column=8).value
            i_val = ws3.cell(row=r, column=9).value
            if not g and not h:
                continue
            if g and "insert row" in str(g):
                continue
            entry = {"row": r}
            if g:
                entry["g"] = str(g)
            if h:
                entry["h"] = str(h)
            if i_val:
                entry["i"] = str(i_val)
            vals = {}
            for c in list(range(23, 35)) + [37, 41, 58, 75]:  # monthly + FY25F, FY26F, FY25B, FY24A
                v = ws3.cell(row=r, column=c).value
                if v is not None and isinstance(v, (int, float)) and v != 0:
                    vals[str(c)] = v
            if vals:
                entry["v"] = vals
            bd_data.append(entry)
        result["pl_breakdown"] = bd_data

        # --- Output CFS ---
        ws4 = wb["Output CFS"]
        CFS_ROWS = list(range(5, 54))
        CFS_COLS = list(range(5, 42))  # E(5) through AO(41), covers all months + FY totals
        cfs_data = {}
        for r in CFS_ROWS:
            rd = {}
            for c in [3] + CFS_COLS:
                v = ws4.cell(row=r, column=c).value
                if v is not None:
                    if isinstance(v, (int, float)):
                        rd[str(c)] = v
                    else:
                        rd[str(c)] = str(v)
            if rd:
                cfs_data[str(r)] = rd
        result["output_cfs"] = cfs_data

        # --- 3b. CFS-breakdown ---
        ws5 = wb["3b. CFS-breakdown"]
        cfs_bd = []
        # Monthly FY26: V(22)-AG(33), FY totals: AN(40)=FY26, AJ(36)=FY25
        CFS_BD_COLS = list(range(22, 34)) + [36, 40]
        for r in range(4, ws5.max_row + 1):
            f_val = ws5.cell(row=r, column=6).value   # F = category
            g_val = ws5.cell(row=r, column=7).value   # G = Platform/Total
            h_val = ws5.cell(row=r, column=8).value   # H = Project
            if not f_val and not g_val:
                continue
            if f_val and "insert row" in str(f_val):
                continue
            entry = {"row": r}
            if f_val: entry["f"] = str(f_val)
            if g_val: entry["g"] = str(g_val)
            if h_val: entry["h"] = str(h_val)
            vals = {}
            for c in CFS_BD_COLS:
                v = ws5.cell(row=r, column=c).value
                if v is not None and isinstance(v, (int, float)) and v != 0:
                    vals[str(c)] = v
            if vals:
                entry["v"] = vals
            cfs_bd.append(entry)
        result["cfs_breakdown"] = cfs_bd

        # --- Restricted Cash Analysis (from 3a. CFS rows 60-66) ---
        ws6 = wb["3a. CFS"]
        rc_analysis = {}
        for r in range(60, 67):
            rd = {}
            label = ws6.cell(row=r, column=6).value  # F = label
            if label:
                rd["label"] = str(label)
            for c in list(range(10, 34)):  # monthly FY25+FY26
                v = ws6.cell(row=r, column=c).value
                if v is not None and isinstance(v, (int, float)):
                    rd[str(c)] = v
            if rd:
                rc_analysis[str(r)] = rd
        result["rc_analysis"] = rc_analysis

        wb.close()

        # Save JSON
        json_path = DISPOSAL_DIR / f"mm_report_{snapshot_name}.json"
        with open(json_path, "w", encoding="utf-8") as f:
            json.dump(result, f)

        return len(pl_data), len(sga_data)
    finally:
        tmp_path.unlink(missing_ok=True)


def get_db():
    db = FeeIncomeDB()
    db.init_db()
    return db

def main():
    st.title("Data Management")
    db = get_db()

    # --- Instructions ---
    with st.expander("How to Upload / Update Data", expanded=False):
        st.markdown("""
        **File naming rule:**
        ```
        Revenue_{YY} Fcst ({N}+{M}).xlsx
        ```
        - `YY` = fiscal year (e.g. 26, 27)
        - `N` = number of actual months, `M` = forecast months (N+M = 12)
        - Examples:
          - `Revenue_26 Fcst (2+10).xlsx` → **FY26 2+10** (Feb close)
          - `Revenue_26 Fcst (3+9).xlsx` → **FY26 3+9** (Mar close)
          - `Revenue_27 Fcst (1+11).xlsx` → **FY27 1+11** (Jan close)
        - Old format also supported: `Revenue_26 Bud and 25 Fcst (2+10).xlsx`

        **How snapshot is recognized:**
        - `Revenue_26` → **FY26**
        - `(2+10)` → **2+10**
        - Combined → **FY26 2+10**

        **Required sheet:**
        - `Summary_new template (작성탭)` — Row 20 = headers, Row 21+ = data

        **Updating data (re-upload):**
        - Upload a file with the same `(N+M)` and `YY` → data is replaced
        - All memos, notes, watch list are **preserved**
        - No need to delete first

        **Deleting a snapshot:**
        - Removes all data **and** associated memos/notes
        - Watch list is not affected (shared across snapshots)
        """)

    # --- Upload ---
    st.header("Upload Excel File")
    uploaded_file = st.file_uploader("Drop your Revenue Excel file here", type=["xlsx"],
        help="Expected: Revenue_26 Fcst (N+M).xlsx")

    if uploaded_file is not None:
        st.caption(f"Selected file: **{uploaded_file.name}**")
        if st.button("Parse and Load", type="primary"):
            with st.spinner("Parsing Excel file..."):
                with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx", prefix="revenue_") as tmp:
                    tmp.write(uploaded_file.getvalue())
                    tmp_path = tmp.name
                try:
                    rows, snapshot = parse_excel_file(tmp_path, original_filename=uploaded_file.name)
                    db.insert_snapshot(snapshot, rows)
                    db.save_snapshot_meta(snapshot, uploaded_file.name)
                    st.cache_data.clear()
                    st.success(f"Loaded **{len(rows):,}** data points for snapshot **{snapshot}**")
                except Exception as e:
                    st.error(f"Error: {e}")
                finally:
                    os.unlink(tmp_path)

    # --- Snapshots ---
    st.header("Snapshots")
    snapshots = db.list_snapshots()
    if not snapshots:
        st.info("No snapshots loaded yet.")
    else:
        for snap in snapshots:
            count = db.query("SELECT COUNT(*) as cnt FROM fee_income WHERE snapshot = ?", (snap,))[0]["cnt"]
            meta = db.get_snapshot_meta(snap)
            filename = meta["filename"] if meta else "-"
            uploaded_at = meta["uploaded_at"] if meta else "-"

            col1, col2, col3, col4 = st.columns([3, 3, 1, 1])
            col1.write(f"**{snap}**")
            col2.write(f"{count:,} rows | `{filename}` | {uploaded_at}")
            if col3.button("Rename", key=f"ren_{snap}"):
                st.session_state[f"renaming_{snap}"] = True
            if col4.button("Delete", key=f"del_{snap}"):
                db.delete_snapshot(snap)
                st.cache_data.clear()
                st.rerun()
            # Rename input row
            if st.session_state.get(f"renaming_{snap}"):
                rc1, rc2 = st.columns([3, 1])
                new_name = rc1.text_input("New name:", value=snap, key=f"rename_input_{snap}")
                if rc2.button("Save", key=f"rename_save_{snap}"):
                    if new_name and new_name != snap:
                        db.rename_snapshot(snap, new_name)
                        st.session_state[f"renaming_{snap}"] = False
                        st.cache_data.clear()
                        st.rerun()
                    else:
                        st.session_state[f"renaming_{snap}"] = False
                        st.rerun()

    # --- Raw Data Viewer ---
    st.header("Raw Data Viewer")
    if snapshots:
        selected_snap = st.selectbox("Snapshot", snapshots)
        limit = st.number_input("Row limit", min_value=10, max_value=10000, value=100)
        raw = db.query("SELECT * FROM fee_income WHERE snapshot = ? LIMIT ?", (selected_snap, limit))
        if raw:
            st.dataframe(pd.DataFrame(raw), use_container_width=True)

    # --- MM Report Upload ---
    st.markdown("---")
    st.header("Monthly Reporting Files")

    st.markdown("""
    **★MM Monthly Reporting** Excel 파일을 snapshot별로 업로드합니다.
    - Snapshot 이름 입력 (예: `2+10`, `3+9`)
    - 같은 snapshot에 다시 업로드하면 파일이 덮어씌워집니다.
    """)

    # Upload
    mm_col1, mm_col2 = st.columns([1, 3])
    with mm_col1:
        mm_snapshot = st.text_input("Snapshot", placeholder="e.g. 2+10", key="mm_snapshot")
    with mm_col2:
        mm_file = st.file_uploader("MM Report Excel file", type=["xlsx"], key="mm_upload")

    if mm_file is not None and mm_snapshot.strip():
        if st.button("Upload MM Report File", type="primary"):
            DISPOSAL_DIR.mkdir(parents=True, exist_ok=True)
            snap = mm_snapshot.strip()
            with st.spinner("Extracting data from Excel (this may take a moment)..."):
                try:
                    pl_count, sga_count = extract_mm_report_json(mm_file.getvalue(), snap)
                    st.success(f"'{mm_file.name}' → **{snap}** snapshot: PL {pl_count} rows, SG&A {sga_count} rows extracted.")
                    st.rerun()
                except Exception as e:
                    st.error(f"Error extracting data: {e}")

    # List existing MM report snapshots
    mm_json_files = sorted(DISPOSAL_DIR.glob("mm_report_*.json"))
    if mm_json_files:
        st.subheader("Uploaded Snapshots")
        for f in mm_json_files:
            snap_name = f.stem.replace("mm_report_", "")
            size_kb = f.stat().st_size / 1024
            mc1, mc2 = st.columns([4, 1])
            mc1.write(f"**{snap_name}** — `{f.name}` ({size_kb:.0f} KB)")
            if mc2.button("Delete", key=f"del_mm_{snap_name}"):
                f.unlink()
                st.rerun()
    else:
        st.info("No MM report files uploaded yet.")

    # --- Disposal Plan Upload ---
    st.markdown("---")
    st.header("Disposal Plan Files")

    st.markdown("""
    **BS and Fund Disposal** Excel 파일을 snapshot별로 업로드합니다.
    - Snapshot 이름을 입력하고 (예: `2+10`, `3+9`) 파일을 업로드하세요.
    - 같은 snapshot에 다시 업로드하면 파일이 덮어씌워집니다.
    """)

    # Upload
    disp_col1, disp_col2 = st.columns([1, 3])
    with disp_col1:
        disp_snapshot = st.text_input("Snapshot", placeholder="e.g. 2+10", key="disp_snapshot")
    with disp_col2:
        disp_file = st.file_uploader("Disposal Excel file", type=["xlsx"], key="disp_upload")

    if disp_file is not None and disp_snapshot.strip():
        if st.button("Upload Disposal File", type="primary"):
            DISPOSAL_DIR.mkdir(parents=True, exist_ok=True)
            target = DISPOSAL_DIR / f"disposal_{disp_snapshot.strip()}.xlsx"
            target.write_bytes(disp_file.getvalue())
            st.success(f"'{disp_file.name}' → **{disp_snapshot.strip()}** snapshot으로 저장되었습니다.")
            st.rerun()

    # List existing disposal snapshots
    disposal_files = sorted(DISPOSAL_DIR.glob("disposal_*.xlsx"))
    if disposal_files:
        st.subheader("Uploaded Snapshots")
        for f in disposal_files:
            snap_name = f.stem.replace("disposal_", "")
            size_kb = f.stat().st_size / 1024
            fc1, fc2 = st.columns([4, 1])
            fc1.write(f"**{snap_name}** — `{f.name}` ({size_kb:.0f} KB)")
            if fc2.button("Delete", key=f"del_disp_{snap_name}"):
                f.unlink()
                st.rerun()
    else:
        st.info("No disposal files uploaded yet.")

main()
