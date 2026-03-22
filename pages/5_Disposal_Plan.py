"""Page 3: Disposal Plan — BS and Fund disposal tracker (Korea only)."""
import streamlit as st
import openpyxl
import pandas as pd
import io
from pathlib import Path
from datetime import datetime
from src.db import FeeIncomeDB


def export_button(df, filename="export.xlsx", key=None):
    buf = io.BytesIO()
    df.to_excel(buf, index=False, engine="openpyxl")
    st.download_button(label="Export to Excel", data=buf.getvalue(),
                        file_name=filename, mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", key=key)

HEADER_COLOR = "#4a5568"
DISPOSAL_DIR = Path(__file__).parent.parent / "data"


def fmt_pct(v):
    if v is None:
        return ""
    return f"{v * 100:.1f}%"


def fmt_m(v):
    """Format number with commas and 2 decimals. Negative in parentheses."""
    if v is None or v == 0:
        return "-"
    if v < 0:
        return f"({abs(v):,.2f})"
    return f"{v:,.2f}"


def fmt_date(v):
    if v is None:
        return ""
    if isinstance(v, datetime):
        return v.strftime("%b-%y")
    return str(v)


def load_bs_disposal_korea(wb):
    """Parse BS Disposal sheet — Korea rows, Forecast only."""
    ws = wb["BS Disposal - Non-China"]
    rows = []
    for row in ws.iter_rows(min_row=6, max_row=ws.max_row, values_only=False):
        region = row[0].value  # A
        if not region or "Korea" not in str(region):
            continue
        status = row[2].value  # C
        if status == "Disposed":
            continue
        rows.append({
            "project": row[1].value,           # B
            "status": status,                  # C
            "esr_stake_bs": row[3].value,      # D - ESR Stake on BS
            "esr_stake_fund": row[4].value,    # E - ESR Stake in Fund
            "gav": row[5].value,               # F - Disposal GAV (a)
            "loan": row[6].value,              # G - Loan (b)
            "other": row[7].value,             # H - Other Asset/Liability (c)
            "nav": row[8].value,               # I - Disposal NAV 100% (d=a+b+c)
            "proceeds": row[9].value,          # J - Attr. Proceeds ESR Portion (e)
            "tax": row[10].value,              # K - Tax Impact (f)
            "adjustments": row[11].value,      # L - Other Adjustments (g)
            "co_invest": row[12].value,        # M - Co-investment (h)
            "capital_recycle": row[13].value,   # N - Capital Recycle (i=e+f+g+h)
            "cash_decon": row[14].value,       # O - Cash-Deconsolidate (j)
            "cash_impact": row[15].value,      # P - Total Cash Impact To ESR (k=i+j)
            "decon_timing": row[16].value,     # Q - Deconsolidation Timing
            "cash_timing": row[17].value,      # R - Cash Collection Timing
        })
    return rows


def load_fund_disposal_korea(wb):
    """Parse Fund Disposal sheet (3rd sheet) — Korea rows only."""
    ws = wb["Fund Disposal - Non-China "]
    rows = []
    for row in ws.iter_rows(min_row=6, max_row=ws.max_row, values_only=False):
        region = row[0].value   # A
        if not region or "Korea" not in str(region):
            continue
        rows.append({
            "project": row[1].value,           # B - Asset Name
            "fund": row[2].value,              # C - Fund Name
            "sale_type": row[3].value,         # D - Sale Type
            "status": row[4].value,            # E - Status
            "esr_stake_cur": row[5].value,     # F - ESR Stake Current
            "esr_stake_tgt": row[6].value,     # G - ESR Stake Target
            "gav": row[7].value,               # H - Disposal GAV (a)
            "loan": row[8].value,              # I - Loan (b)
            "other": row[9].value,             # J - Other Asset/Liability (c)
            "nav": row[10].value,              # K - Disposal NAV 100% (d=a+b+c)
            "proceeds": row[11].value,         # L - Attr. Proceeds ESR Portion (e)
            "tax": row[12].value,              # M - Tax Impact (f)
            "adjustments": row[13].value,      # N - Other Adjustments (g)
            "co_invest": row[14].value,        # O - Co-investment (h)
            "capital_recycle": row[15].value,   # P - Capital Recycle (i=e+f+g+h)
            "cash_decon": row[16].value,       # Q - Cash-Deconsolidate (j)
            "cash_impact": row[17].value,      # R - Total Cash Impact To ESR (k=i+j)
            "decon_timing": row[18].value,     # S - Deconsolidation Timing
            "cash_timing": row[19].value,      # T - Cash Collection Timing
        })
    return rows


# Column keys for summing totals (shared by BS and Fund)
SUM_COLS = ["gav", "loan", "other", "nav", "proceeds", "tax", "adjustments",
            "co_invest", "capital_recycle", "cash_decon", "cash_impact"]


def render_bs_table(rows):
    TH = "padding:6px 8px; border:1px solid #cbd5e0;"
    html = f"""<table style="border-collapse:collapse; width:100%; font-size:12px; font-family:Calibri,sans-serif;">
    <thead>
    <tr style="background:{HEADER_COLOR}; color:white; font-weight:bold; text-align:center;">
        <th style="{TH} text-align:left;" rowspan="2">Project name</th>
        <th style="{TH}" rowspan="2">Status</th>
        <th style="{TH}" colspan="2">ESR Stake</th>
        <th style="{TH}" colspan="4">Disposal NAV</th>
        <th style="{TH}" colspan="5">Capital Recycle</th>
        <th style="{TH}" rowspan="2">Cash-Decon<br>(if applicable)</th>
        <th style="{TH}" rowspan="2">Total Cash Impact<br>To ESR</th>
        <th style="{TH}" colspan="2">Timing</th>
    </tr>
    <tr style="background:{HEADER_COLOR}; color:white; font-weight:bold; text-align:center; font-size:11px;">
        <th style="{TH}">on BS</th>
        <th style="{TH}">in Fund</th>
        <th style="{TH}">GAV</th>
        <th style="{TH}">Loan</th>
        <th style="{TH}">Other<br>Asset/Liability</th>
        <th style="{TH}">NAV (100%)</th>
        <th style="{TH}">Attr. Proceeds<br>(ESR Portion)</th>
        <th style="{TH}">Tax<br>Impact</th>
        <th style="{TH}">Other<br>Adjustments</th>
        <th style="{TH}">Co-investment</th>
        <th style="{TH}">Capital<br>Recycle</th>
        <th style="{TH}">Deconsolidation</th>
        <th style="{TH}">Cash Collection</th>
    </tr>
    </thead><tbody>"""

    totals = {k: 0 for k in SUM_COLS}
    TD = "padding:5px 8px; border:1px solid #cbd5e0;"
    for i, r in enumerate(rows):
        bg = "#f7fafc" if i % 2 == 0 else "#ffffff"
        for k in SUM_COLS:
            totals[k] += (r[k] or 0)
        html += f"""<tr style="background:{bg};">
            <td style="{TD} font-weight:bold;">{r['project'] or ''}</td>
            <td style="{TD} text-align:center; color:#2b6cb0;">{r['status'] or ''}</td>
            <td style="{TD} text-align:right;">{fmt_pct(r['esr_stake_bs'])}</td>
            <td style="{TD} text-align:right;">{fmt_pct(r['esr_stake_fund'])}</td>
            <td style="{TD} text-align:right;">{fmt_m(r['gav'])}</td>
            <td style="{TD} text-align:right;">{fmt_m(r['loan'])}</td>
            <td style="{TD} text-align:right;">{fmt_m(r['other'])}</td>
            <td style="{TD} text-align:right; font-weight:bold;">{fmt_m(r['nav'])}</td>
            <td style="{TD} text-align:right;">{fmt_m(r['proceeds'])}</td>
            <td style="{TD} text-align:right;">{fmt_m(r['tax'])}</td>
            <td style="{TD} text-align:right;">{fmt_m(r['adjustments'])}</td>
            <td style="{TD} text-align:right;">{fmt_m(r['co_invest'])}</td>
            <td style="{TD} text-align:right;">{fmt_m(r['capital_recycle'])}</td>
            <td style="{TD} text-align:right;">{fmt_m(r['cash_decon'])}</td>
            <td style="{TD} text-align:right; font-weight:bold;">{fmt_m(r['cash_impact'])}</td>
            <td style="{TD} text-align:center;">{fmt_date(r['decon_timing'])}</td>
            <td style="{TD} text-align:center;">{fmt_date(r['cash_timing'])}</td>
        </tr>"""

    html += f"""<tr style="background:{HEADER_COLOR}; color:white; font-weight:bold;">
        <td style="{TD}" colspan="4">Grand Total</td>
        <td style="{TD} text-align:right;">{fmt_m(totals['gav'])}</td>
        <td style="{TD} text-align:right;">{fmt_m(totals['loan'])}</td>
        <td style="{TD} text-align:right;">{fmt_m(totals['other'])}</td>
        <td style="{TD} text-align:right;">{fmt_m(totals['nav'])}</td>
        <td style="{TD} text-align:right;">{fmt_m(totals['proceeds'])}</td>
        <td style="{TD} text-align:right;">{fmt_m(totals['tax'])}</td>
        <td style="{TD} text-align:right;">{fmt_m(totals['adjustments'])}</td>
        <td style="{TD} text-align:right;">{fmt_m(totals['co_invest'])}</td>
        <td style="{TD} text-align:right;">{fmt_m(totals['capital_recycle'])}</td>
        <td style="{TD} text-align:right;">{fmt_m(totals['cash_decon'])}</td>
        <td style="{TD} text-align:right;">{fmt_m(totals['cash_impact'])}</td>
        <td style="{TD}" colspan="2"></td>
    </tr>"""
    html += "</tbody></table>"
    return html, totals["cash_impact"]


def render_fund_table(rows):
    TH = "padding:6px 8px; border:1px solid #cbd5e0;"
    html = f"""<table style="border-collapse:collapse; width:100%; font-size:12px; font-family:Calibri,sans-serif;">
    <thead>
    <tr style="background:{HEADER_COLOR}; color:white; font-weight:bold; text-align:center;">
        <th style="{TH} text-align:left;" rowspan="2">Asset Name</th>
        <th style="{TH} text-align:left;" rowspan="2">Fund Name</th>
        <th style="{TH}" rowspan="2">Sale Type</th>
        <th style="{TH}" rowspan="2">Status</th>
        <th style="{TH}" colspan="2">ESR Stake</th>
        <th style="{TH}" colspan="4">Disposal NAV</th>
        <th style="{TH}" colspan="5">Capital Recycle</th>
        <th style="{TH}" rowspan="2">Cash-Decon<br>(if applicable)</th>
        <th style="{TH}" rowspan="2">Total Cash Impact<br>To ESR</th>
        <th style="{TH}" colspan="2">Timing</th>
    </tr>
    <tr style="background:{HEADER_COLOR}; color:white; font-weight:bold; text-align:center; font-size:11px;">
        <th style="{TH}">Current</th>
        <th style="{TH}">Target</th>
        <th style="{TH}">GAV</th>
        <th style="{TH}">Loan</th>
        <th style="{TH}">Other<br>Asset/Liability</th>
        <th style="{TH}">NAV (100%)</th>
        <th style="{TH}">Attr. Proceeds<br>(ESR Portion)</th>
        <th style="{TH}">Tax<br>Impact</th>
        <th style="{TH}">Other<br>Adjustments</th>
        <th style="{TH}">Co-investment</th>
        <th style="{TH}">Capital<br>Recycle</th>
        <th style="{TH}">Deconsolidation</th>
        <th style="{TH}">Cash Collection</th>
    </tr>
    </thead><tbody>"""

    totals = {k: 0 for k in SUM_COLS}
    TD = "padding:5px 8px; border:1px solid #cbd5e0;"
    for i, r in enumerate(rows):
        bg = "#f7fafc" if i % 2 == 0 else "#ffffff"
        status = r["status"] or ""
        status_color = "#718096" if status == "Disposed" else "#2b6cb0"
        for k in SUM_COLS:
            totals[k] += (r[k] or 0)
        html += f"""<tr style="background:{bg};">
            <td style="{TD} font-weight:bold;">{r['project'] or ''}</td>
            <td style="{TD}">{r['fund'] or ''}</td>
            <td style="{TD} text-align:center;">{r['sale_type'] or ''}</td>
            <td style="{TD} text-align:center; color:{status_color};">{status}</td>
            <td style="{TD} text-align:right;">{fmt_pct(r['esr_stake_cur'])}</td>
            <td style="{TD} text-align:right;">{fmt_pct(r['esr_stake_tgt'])}</td>
            <td style="{TD} text-align:right;">{fmt_m(r['gav'])}</td>
            <td style="{TD} text-align:right;">{fmt_m(r['loan'])}</td>
            <td style="{TD} text-align:right;">{fmt_m(r['other'])}</td>
            <td style="{TD} text-align:right; font-weight:bold;">{fmt_m(r['nav'])}</td>
            <td style="{TD} text-align:right;">{fmt_m(r['proceeds'])}</td>
            <td style="{TD} text-align:right;">{fmt_m(r['tax'])}</td>
            <td style="{TD} text-align:right;">{fmt_m(r['adjustments'])}</td>
            <td style="{TD} text-align:right;">{fmt_m(r['co_invest'])}</td>
            <td style="{TD} text-align:right;">{fmt_m(r['capital_recycle'])}</td>
            <td style="{TD} text-align:right;">{fmt_m(r['cash_decon'])}</td>
            <td style="{TD} text-align:right; font-weight:bold;">{fmt_m(r['cash_impact'])}</td>
            <td style="{TD} text-align:center;">{fmt_date(r['decon_timing'])}</td>
            <td style="{TD} text-align:center;">{fmt_date(r['cash_timing'])}</td>
        </tr>"""

    html += f"""<tr style="background:{HEADER_COLOR}; color:white; font-weight:bold;">
        <td style="{TD}" colspan="6">Grand Total</td>
        <td style="{TD} text-align:right;">{fmt_m(totals['gav'])}</td>
        <td style="{TD} text-align:right;">{fmt_m(totals['loan'])}</td>
        <td style="{TD} text-align:right;">{fmt_m(totals['other'])}</td>
        <td style="{TD} text-align:right;">{fmt_m(totals['nav'])}</td>
        <td style="{TD} text-align:right;">{fmt_m(totals['proceeds'])}</td>
        <td style="{TD} text-align:right;">{fmt_m(totals['tax'])}</td>
        <td style="{TD} text-align:right;">{fmt_m(totals['adjustments'])}</td>
        <td style="{TD} text-align:right;">{fmt_m(totals['co_invest'])}</td>
        <td style="{TD} text-align:right;">{fmt_m(totals['capital_recycle'])}</td>
        <td style="{TD} text-align:right;">{fmt_m(totals['cash_decon'])}</td>
        <td style="{TD} text-align:right;">{fmt_m(totals['cash_impact'])}</td>
        <td style="{TD}" colspan="2"></td>
    </tr>"""
    html += "</tbody></table>"
    return html, totals["cash_impact"]


def render_comments_display(table_key, projects, db):
    """Show saved comments as a styled HTML table."""
    saved = db.get_drivers("disposal", table_key)
    has_comments = any(saved.get(p) for p in projects)
    if not has_comments:
        return
    TD = "padding:5px 8px; border:1px solid #cbd5e0;"
    html = f"""<table style="border-collapse:collapse; width:100%; font-size:12px; font-family:Calibri,sans-serif; margin-top:8px;">
    <thead><tr style="background:{HEADER_COLOR}; color:white; font-weight:bold;">
        <th style="{TD} text-align:left; width:20%;">Project</th>
        <th style="{TD} text-align:left;">Comment</th>
    </tr></thead><tbody>"""
    for i, proj in enumerate(projects):
        comment = saved.get(proj, "")
        if not comment:
            continue
        bg = "#f7fafc" if i % 2 == 0 else "#ffffff"
        html += f"""<tr style="background:{bg};">
            <td style="{TD} font-weight:bold;">{proj}</td>
            <td style="{TD}">{comment}</td>
        </tr>"""
    html += "</tbody></table>"
    st.markdown(html, unsafe_allow_html=True)


def render_comments_editor(table_key, projects, db):
    saved = db.get_drivers("disposal", table_key)
    with st.expander("Edit Comments"):
        updated = {}
        for proj in projects:
            updated[proj] = st.text_area(
                proj, value=saved.get(proj, ""), height=68,
                key=f"comment_{table_key}_{proj}",
            )
        if st.button("Save Comments", key=f"save_comments_{table_key}"):
            db.save_drivers("disposal", table_key, updated)
            st.success("Saved!")
            st.rerun()


def main():
    st.title("Disposal Plan — Korea")

    db = FeeIncomeDB()
    db.init_db()

    # --- Snapshot selector ---
    disposal_files = sorted(DISPOSAL_DIR.glob("disposal_*.xlsx"), reverse=True)
    if not disposal_files:
        st.warning("No disposal data found. Go to Data Management to upload.")
        return

    snap_names = [f.stem.replace("disposal_", "") for f in disposal_files]
    selected = st.selectbox("Snapshot", snap_names)
    data_file = DISPOSAL_DIR / f"disposal_{selected}.xlsx"

    wb = openpyxl.load_workbook(str(data_file), data_only=True)

    bs_rows = load_bs_disposal_korea(wb)
    fund_rows = load_fund_disposal_korea(wb)

    # --- Summary Cards ---
    bs_cash_total = sum((r["cash_impact"] or 0) for r in bs_rows)
    fund_cash_total = sum((r["cash_impact"] or 0) for r in fund_rows)
    total_cash = bs_cash_total + fund_cash_total

    col1, col2, col3 = st.columns(3)
    with col1:
        st.markdown(f"""
        <div style="background:#f7fafc; border-radius:6px; padding:14px 16px; border-top:3px solid {HEADER_COLOR};">
            <div style="font-size:12px; color:#718096; text-transform:uppercase;">Total Cash Impact (Korea)</div>
            <div style="font-size:22px; font-weight:bold; color:#2d3748;">${total_cash:,.2f}M</div>
        </div>""", unsafe_allow_html=True)
    with col2:
        st.markdown(f"""
        <div style="background:#f7fafc; border-radius:6px; padding:14px 16px; border-top:3px solid #2b6cb0;">
            <div style="font-size:12px; color:#718096; text-transform:uppercase;">BS Disposal Pipeline</div>
            <div style="font-size:22px; font-weight:bold; color:#2b6cb0;">{len(bs_rows)} projects</div>
        </div>""", unsafe_allow_html=True)
    with col3:
        st.markdown(f"""
        <div style="background:#f7fafc; border-radius:6px; padding:14px 16px; border-top:3px solid #2b6cb0;">
            <div style="font-size:12px; color:#718096; text-transform:uppercase;">Fund Disposal Pipeline</div>
            <div style="font-size:22px; font-weight:bold; color:#2b6cb0;">{len(fund_rows)} projects</div>
        </div>""", unsafe_allow_html=True)

    st.markdown("")

    # --- BS Disposal ---
    st.subheader("BS Disposal — Korea")
    if bs_rows:
        bs_html, _ = render_bs_table(bs_rows)
        st.markdown(bs_html, unsafe_allow_html=True)
        st.caption("Unit: USD millions")
        exp_rows = []
        for r in bs_rows:
            exp_rows.append({
                "Project": r["project"] or "",
                "Status": r["status"] or "",
                "ESR Stake on BS": r["esr_stake_bs"],
                "ESR Stake in Fund": r["esr_stake_fund"],
                **{col.replace("_"," ").title() + " (USD)": round((r[col] or 0) * 1_000_000, 2) for col in SUM_COLS},
                "Decon Timing": fmt_date(r["decon_timing"]),
                "Cash Timing": fmt_date(r["cash_timing"]),
            })
        export_button(pd.DataFrame(exp_rows), "bs_disposal_korea.xlsx", key="export_bs_disposal")
    else:
        st.info("No BS Disposal forecast data for Korea.")
    bs_projects = [r["project"] for r in bs_rows]
    render_comments_display("bs_korea", bs_projects, db)
    render_comments_editor("bs_korea", bs_projects, db)

    st.markdown("---")

    # --- Fund Disposal ---
    st.subheader("Fund Disposal — Korea")
    if fund_rows:
        fund_html, _ = render_fund_table(fund_rows)
        st.markdown(fund_html, unsafe_allow_html=True)
        st.caption("Unit: USD millions")
        exp_rows = []
        for r in fund_rows:
            exp_rows.append({
                "Asset": r["project"] or "",
                "Fund": r["fund"] or "",
                "Sale Type": r["sale_type"] or "",
                "Status": r["status"] or "",
                "ESR Stake Current": r["esr_stake_cur"],
                "ESR Stake Target": r["esr_stake_tgt"],
                **{col.replace("_"," ").title() + " (USD)": round((r[col] or 0) * 1_000_000, 2) for col in SUM_COLS},
                "Decon Timing": fmt_date(r["decon_timing"]),
                "Cash Timing": fmt_date(r["cash_timing"]),
            })
        export_button(pd.DataFrame(exp_rows), "fund_disposal_korea.xlsx", key="export_fund_disposal")
    else:
        st.info("No Fund Disposal data for Korea.")
    fund_projects = [r["project"] for r in fund_rows]
    render_comments_display("fund_korea", fund_projects, db)
    render_comments_editor("fund_korea", fund_projects, db)


main()
